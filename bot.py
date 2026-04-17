from flask import Flask, render_template, request, jsonify
import openpyxl
from groq import Groq
import json
import os
from docx import Document
from dotenv import load_dotenv
from utils.faq_manager import load_uploaded_faqs, save_faq_json, add_faq, delete_faq, edit_faq, get_faq_by_id, extract_faqs_from_text

# Load .env file for local development (harmless in production)
load_dotenv()

app = Flask(__name__)

# Initialize Groq client with error handling
client = None
try:
    api_key = os.getenv("GROQ_API_KEY")
    print(f"DEBUG: API Key from env: {'SET' if api_key else 'NOT SET'}")
    if api_key:
        print(f"DEBUG: API Key length: {len(api_key)}")
        print(f"DEBUG: API Key starts with: {api_key[:10]}")

    if not api_key:
        raise ValueError("GROQ_API_KEY environment variable not set")

    print("DEBUG: Attempting to initialize Groq client...")
    client = Groq(api_key=api_key)
    print("✓ Groq API available")
    print(f"✓ API Key found (length: {len(api_key)})")
except Exception as e:
    print(f"✗ Groq initialization failed: {e}")
    print(f"✗ Error type: {type(e).__name__}")
    print("Set GROQ_API_KEY environment variable to use the bot")

# Load Excel templates
def load_templates():
    """Load Bursary templates organized by subcategories"""
    try:
        wb = openpyxl.load_workbook('data/TEMPLATE_Replies Catergorisation.xlsx')
        ws = wb['Bursary']

        templates = {}

        # Row 2 contains category names
        category_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        # Extract categories and their column indices
        categories = {}
        for col_idx, category in enumerate(category_row):
            if category:
                categories[col_idx] = category

        print(f"✓ Found {len(categories)} categories: {list(categories.values())}")

        # Initialize category lists
        for cat in categories.values():
            templates[cat] = []

        # Extract templates from rows 3 onwards
        template_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            for col_idx, category in categories.items():
                cell_value = row[col_idx]
                if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 20:
                    templates[category].append(cell_value.strip())
                    template_count += 1

        print(f"✓ Loaded {template_count} templates")
        return templates

    except Exception as e:
        print(f"✗ Error loading templates: {e}")
        return {}

# Initialize templates
TEMPLATES = load_templates()

# Load uploaded FAQs and add to templates
uploaded_faqs = load_uploaded_faqs()

# Flatten all templates for easier searching with category info
ALL_TEMPLATES = []
for category, texts in TEMPLATES.items():
    for text in texts:
        ALL_TEMPLATES.append({
            'category': category,
            'text': text
        })

# Add uploaded FAQs to template pool
for faq in uploaded_faqs:
    ALL_TEMPLATES.append({
        'category': faq.get('category', 'User Uploaded'),
        'text': faq['text'],
        'id': faq['id']
    })

from difflib import SequenceMatcher

def find_best_template(faq):
    """Use Groq to find the best matching Bursary template for a FAQ"""

    # Calculate similarity scores for each template
    template_scores = []
    faq_lower = faq.lower()

    for template in ALL_TEMPLATES:
        template_text = template['text'].lower()
        # Simple similarity: count matching words
        faq_words = set(faq_lower.split())
        template_words = set(template_text.split())

        # Calculate overlap
        overlap = len(faq_words & template_words)

        template_scores.append({
            'template': template,
            'score': overlap
        })

    # Sort by relevance score (highest first) and take top 20
    sorted_templates = sorted(template_scores, key=lambda x: x['score'], reverse=True)[:20]

    # Create a formatted list of top templates
    templates_text = "\n\n---\n\n".join([
        f"[{t['template']['category']}]\n{t['template']['text']}"
        for t in sorted_templates
    ])

    response = client.chat.completions.create(
        model="llama-3.1-8b-instant",
        messages=[
            {
                "role": "user",
                "content": f"""You are an email response assistant for Ngee Ann Polytechnic student Bursary inquiries.

A student sent this inquiry:
"{faq}"

Below are reference templates from similar past inquiries:

REFERENCE TEMPLATES:
{templates_text}

STRICT INSTRUCTIONS:
1. Carefully review ALL the reference templates above
2. ONLY provide information that is explicitly mentioned or clearly implied in the templates
3. Generate an ORIGINAL response that:
   - Directly addresses the student's specific inquiry
   - Uses ONLY accurate information found in the templates
   - Maintains the professional tone and style of the templates
   - Does NOT copy any template verbatim
4. If the answer to the student's inquiry is NOT found in the templates, say: "I don't have enough information to answer this inquiry. Please escalate to the appropriate department."
5. Every fact or policy mentioned MUST be backed by evidence from the templates above

Write your response:"""
            }
        ],
        temperature=0.3
    )

    return response.choices[0].message.content

@app.route('/debug/env', methods=['GET'])
def debug_env():
    """Debug endpoint - shows environment variables and template status"""
    return jsonify({
        'GROQ_API_KEY': 'SET' if os.getenv('GROQ_API_KEY') else 'NOT SET',
        'PORT': os.getenv('PORT', 'not set'),
        'templates_count': len(ALL_TEMPLATES),
        'excel_categories': list(TEMPLATES.keys()),
        'uploaded_faqs_count': len(load_uploaded_faqs()),
        'sample_templates': [t['text'][:50] for t in ALL_TEMPLATES[:3]]
    })

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/templates')
def templates():
    return render_template('templates.html')

@app.route('/api/templates', methods=['GET'])
def get_templates():
    """Get all Bursary templates organized by subcategories"""
    return jsonify({
        'templates': ALL_TEMPLATES,
        'categories': list(TEMPLATES.keys()),
        'total': len(ALL_TEMPLATES)
    })

@app.route('/api/upload-faq', methods=['POST'])
def upload_faq():
    """Upload a Word document and extract FAQs"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']

        if not file.filename.endswith('.docx'):
            return jsonify({'error': 'Only .docx files allowed'}), 400

        # Extract text from Word document
        doc = Document(file)
        text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])

        if not text.strip():
            return jsonify({'error': 'No text found in document'}), 400

        # Extract FAQs from text
        extracted_faqs = extract_faqs_from_text(text)

        if not extracted_faqs:
            # If no Q&A pattern found, treat entire text as one FAQ
            extracted_faqs = [{'q': 'Extracted Content', 'a': text[:200]}]

        return jsonify({
            'success': True,
            'extracted': extracted_faqs,
            'raw_text': text,
            'filename': file.filename
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-faq/confirm', methods=['POST'])
def confirm_upload_faq():
    """Save extracted FAQs to JSON storage - each FAQ separately"""
    try:
        data = request.json
        faqs_data = data.get('faqs', [])  # Array of FAQ dicts
        filename = data.get('filename', 'Unknown')

        if not faqs_data or len(faqs_data) == 0:
            return jsonify({'error': 'No FAQs to save'}), 400

        saved_faqs = []
        for faq_dict in faqs_data:
            new_faq = add_faq(faq_dict, filename)
            saved_faqs.append(new_faq)

        # Reload templates to include new FAQs
        global ALL_TEMPLATES
        uploaded_faqs = load_uploaded_faqs()
        ALL_TEMPLATES = []
        for category, texts in TEMPLATES.items():
            for text in texts:
                ALL_TEMPLATES.append({'category': category, 'text': text})
        for faq in uploaded_faqs:
            # Use question+answer as text for template pool
            faq_text = f"Q: {faq.get('question', 'FAQ')}\nA: {faq.get('answer', faq.get('text', ''))}"
            ALL_TEMPLATES.append({
                'category': faq.get('category', 'User Uploaded'),
                'text': faq_text,
                'id': faq['id']
            })

        return jsonify({
            'success': True,
            'message': f'Saved {len(saved_faqs)} FAQs successfully',
            'faqs': saved_faqs
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/uploaded-faqs', methods=['GET'])
def get_uploaded_faqs():
    """Get all uploaded FAQs"""
    try:
        faqs = load_uploaded_faqs()
        return jsonify({
            'success': True,
            'faqs': faqs,
            'total': len(faqs)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/uploaded-faqs/<faq_id>', methods=['DELETE'])
def delete_uploaded_faq(faq_id):
    """Delete an uploaded FAQ"""
    try:
        delete_faq(faq_id)

        # Reload templates
        global ALL_TEMPLATES
        uploaded_faqs = load_uploaded_faqs()
        ALL_TEMPLATES = []
        for category, texts in TEMPLATES.items():
            for text in texts:
                ALL_TEMPLATES.append({'category': category, 'text': text})
        for faq in uploaded_faqs:
            ALL_TEMPLATES.append({
                'category': faq.get('category', 'User Uploaded'),
                'text': faq['text'],
                'id': faq['id']
            })

        return jsonify({
            'success': True,
            'message': 'FAQ deleted successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/uploaded-faqs/<faq_id>', methods=['PUT'])
def update_uploaded_faq(faq_id):
    """Update an uploaded FAQ"""
    try:
        data = request.json
        new_text = data.get('text', '').strip()

        if not new_text:
            return jsonify({'error': 'FAQ text cannot be empty'}), 400

        edit_faq(faq_id, new_text)

        # Reload templates
        global ALL_TEMPLATES
        uploaded_faqs = load_uploaded_faqs()
        ALL_TEMPLATES = []
        for category, texts in TEMPLATES.items():
            for text in texts:
                ALL_TEMPLATES.append({'category': category, 'text': text})
        for faq in uploaded_faqs:
            ALL_TEMPLATES.append({
                'category': faq.get('category', 'User Uploaded'),
                'text': faq['text'],
                'id': faq['id']
            })

        return jsonify({
            'success': True,
            'message': 'FAQ updated successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-reply', methods=['POST'])
def generate_reply():
    """Generate a reply for the given FAQ"""
    if not client:
        return jsonify({'error': 'Groq API not configured. Set GROQ_API_KEY environment variable.'}), 500

    data = request.json
    faq = data.get('faq', '').strip()

    if not faq:
        return jsonify({'error': 'FAQ cannot be empty'}), 400

    try:
        reply = find_best_template(faq)
        return jsonify({'reply': reply})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
